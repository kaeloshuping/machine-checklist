from openpyxl import load_workbook
from data import printers

'''
This class takes the last 3 digits of a machine serial number 
and retrieves the meter reading readings in either the
current/ previous/ difference meter readings column and returns 
it in a dictionary form
'''


class GetReadings:

    def __init__(self, serial_number):
        self.serial_number = serial_number
        # load excel file
        self.workbook = load_workbook(filename="./machine_list.xlsx")

        # open current worksheet "machines"
        self.worksheet = self.workbook["machines"]

        if serial_number in printers["B&W"]:
            self.machine_type = "B&W"

        else:
            self.machine_type = "COLOUR"

    # This method takes a column number for either CURRENT or PREVIOUS or DIFFERENCE columns and
    # returns the meter reading in a dictionary format
    def get_reading_dict(self, column_number):
        # iterate through machines column and store cell object in machine_name
        for row in self.worksheet.iter_rows():
            machine_name = row[0]

            # check whether user input is equal to one of the values in "MACHINE_NAME"
            if printers[self.machine_type][self.serial_number] == machine_name.value:

                # if user input is equal to one of the values in "cell", retrieve value of "current meter reading"
                meter_reading = self.worksheet.cell(row=machine_name.row, column=column_number).value

                # convert meter_reading into list of strings
                readings_string_list = meter_reading.replace("-", ":").split()

                # initialize empty dictionary
                reading_dict = {}

                # loop through list of strings
                for string in readings_string_list:

                    # convert list of strings to "key", "value" pairs for a dictionary
                    key, *value = string.split(":")

                    # add "key", "value" pairs to empty dictionary "colour_reading_dict"
                    reading_dict[key] = int("".join(value))
                return reading_dict
